
import os
import json
import math
from dataclasses import dataclass, asdict
from typing import Dict, Any, Optional, Tuple, List

import csv
import openpyxl
from dotenv import load_dotenv

from telegram import (
    Update, InlineKeyboardButton, InlineKeyboardMarkup, InputFile
)
from telegram.constants import ParseMode
from telegram.ext import (
    Application, CommandHandler, CallbackQueryHandler, MessageHandler,
    ContextTypes, ConversationHandler, filters
)

# ---------------------------
# Config / Persistence
# ---------------------------
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE_DIR, "data")
PREDICTOR_XLSX = os.path.join(BASE_DIR, "Predictor.xlsx")

LECITHIN_KEY = "lecithin_logs_v1"
SHIFT_KEY = "gum_shift_logs_v1"

def _user_file(chat_id: int) -> str:
    return os.path.join(DATA_DIR, f"{chat_id}.json")

def load_user_data(chat_id: int) -> Dict[str, Any]:
    path = _user_file(chat_id)
    if not os.path.exists(path):
        return {LECITHIN_KEY: {}, SHIFT_KEY: {}}
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)
    data.setdefault(LECITHIN_KEY, {})
    data.setdefault(SHIFT_KEY, {})
    return data

def save_user_data(chat_id: int, data: Dict[str, Any]) -> None:
    os.makedirs(DATA_DIR, exist_ok=True)
    with open(_user_file(chat_id), "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def _admin_ids() -> set[int]:
    """
    خواندن لیست ادمین‌ها از ENV.
    مقدار پیشنهادی: ADMIN_CHAT_ID="12345,67890"
    """
    raw = os.getenv("ADMIN_CHAT_ID", "").strip()
    if not raw:
        return set()
    ids: set[int] = set()
    for part in raw.replace(";", ",").split(","):
        part = part.strip()
        if not part:
            continue
        try:
            ids.add(int(part))
        except ValueError:
            continue
    return ids

async def myid(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """نمایش Chat ID برای ست کردن ADMIN_CHAT_ID در Railway."""
    chat_id = update.effective_chat.id if update.effective_chat else None
    user_id = update.effective_user.id if update.effective_user else None
    msg = f"""🆔 Chat ID: {chat_id}
👤 User ID: {user_id}

برای فعال‌سازی آمار مدیریتی، در Railway یک متغیر بسازید:
ADMIN_CHAT_ID="{chat_id}"
"""
    await update.message.reply_text(msg)

async def admin_stats(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """
    آمار کلی استفاده از ربات (فقط ادمین).
    - تعداد کاربران (براساس تعداد فایل‌های data/*.json)
    - تعداد ثبت‌های لسیتین و ارزیابی
    """
    chat_id = update.effective_chat.id if update.effective_chat else 0
    admins = _admin_ids()

    if not admins:
        msg = """⚠️ آمار مدیریتی فعال نیست.
ابتدا در Railway یک متغیر ENV بسازید:
ADMIN_CHAT_ID="CHAT_ID"

برای گرفتن Chat ID خودتان دستور /myid را بزنید.
"""
        await update.message.reply_text(msg)
        return

    if chat_id not in admins:
        await update.message.reply_text("⛔️ دسترسی ندارید.")
        return

    os.makedirs(DATA_DIR, exist_ok=True)
    user_files = [fn for fn in os.listdir(DATA_DIR) if fn.endswith(".json")]

    users = len(user_files)
    lecithin_count = 0
    shift_count = 0
    sites: dict[str, int] = {"Semnan": 0, "Kermanshah": 0}

    for fn in user_files:
        try:
            with open(os.path.join(DATA_DIR, fn), "r", encoding="utf-8") as f:
                data = json.load(f)
        except Exception:
            continue

        lec = data.get(LECITHIN_KEY, {}) if isinstance(data, dict) else {}
        shf = data.get(SHIFT_KEY, {}) if isinstance(data, dict) else {}

        if isinstance(lec, dict):
            lecithin_count += len(lec)
            for _k, rec in lec.items():
                if isinstance(rec, dict):
                    s = rec.get("site")
                    if s in sites:
                        sites[s] += 1

        if isinstance(shf, dict):
            shift_count += len(shf)
            for _k, rec in shf.items():
                if isinstance(rec, dict):
                    s = rec.get("site")
                    if s in sites:
                        sites[s] += 1

    msg = f"""📊 *آمار مدیریتی ربات*

👥 تعداد کاربران: *{users}*
🧪 تعداد ثبت‌های پیش‌بینی لسیتین: *{lecithin_count}*
👷 تعداد ثبت‌های ارزیابی عملکرد: *{shift_count}*

🏭 ثبت‌ها به تفکیک سایت (تقریبی):
• سمنان: *{sites['Semnan']}*
• کرمانشاه: *{sites['Kermanshah']}*
"""
    await update.message.reply_text(msg, parse_mode=ParseMode.MARKDOWN)

# ---------------------------
# Excel -> grid cache
# ---------------------------
def sheet_to_matrix(ws) -> List[List[Any]]:
    max_row = ws.max_row
    max_col = ws.max_column
    rows: List[List[Any]] = []
    for r in range(1, max_row + 1):
        row = []
        for c in range(1, max_col + 1):
            row.append(ws.cell(row=r, column=c).value)
        # keep row length fixed (like sheet_to_json header:1)
        rows.append(row)
    return rows

class PredictorData:
    def __init__(self, xlsx_path: str):
        wb = openpyxl.load_workbook(xlsx_path, data_only=False)
        self.sheets: Dict[str, List[List[Any]]] = {}
        for name in wb.sheetnames:
            self.sheets[name] = sheet_to_matrix(wb[name])

PRED = PredictorData(PREDICTOR_XLSX)

# ---------------------------
# Core math (same as web app)
# ---------------------------
def lerp(x: float, x0: float, x1: float, y0: float, y1: float) -> float:
    if x0 == x1:
        return float(y0)
    return float(y0) + (x - x0) * (float(y1) - float(y0)) / (x1 - x0)

def find_indices(axis: List[float], val: float) -> Tuple[int, int]:
    if val <= axis[0]:
        return (0, 0)
    if val >= axis[-1]:
        return (len(axis) - 1, len(axis) - 1)
    for i in range(len(axis) - 1):
        if axis[i] <= val <= axis[i + 1]:
            return (i, i + 1)
    return (len(axis) - 1, len(axis) - 1)

def trilinear_interpolate(grid: List[List[Any]], ffa: float, ton: float, target_hours: float) -> float:
    # grid[0][1] is B1, but our matrix is 0-based; B is index 1
    try:
        ffa_count = int(grid[0][1])
        ton_count = int(grid[1][1])
        hour_count = int(grid[2][1])
    except Exception:
        raise ValueError("داده‌های ابعادی محور در MonoGrid نامعتبر است.")

    # axes start at column E => index 4
    def to_float(x):
        try:
            return float(x)
        except Exception:
            return float("nan")

    ffa_axis = [to_float(v) for v in grid[0][4:4 + ffa_count]]
    ton_axis = [to_float(v) for v in grid[1][4:4 + ton_count]]
    hour_axis = [to_float(v) for v in grid[2][4:4 + hour_count]]

    if any(math.isnan(v) for v in (ffa_axis + ton_axis + hour_axis)):
        raise ValueError("محورها در MonoGrid نامعتبر است.")

    i0, i1 = find_indices(ffa_axis, ffa)
    j0, j1 = find_indices(ton_axis, ton)
    k0, k1 = find_indices(hour_axis, target_hours)

    n_ton = len(ton_axis)

    def get_val(i: int, j: int, k: int) -> float:
        # JS:
        # blockStart = 5 + i*(Nton+2)
        # rowIndex = blockStart + j
        # colIndex = 1 + k
        block_start = 5 + i * (n_ton + 2)
        row_index = block_start + j
        col_index = 1 + k  # B=1, C=2, ...
        try:
            v = grid[row_index][col_index]
        except Exception:
            v = None
        try:
            fv = float(v)
        except Exception:
            raise ValueError(f"شکاف داده در MonoGrid: بلوک {i}، ردیف {j}، ستون {k}")
        if math.isnan(fv):
            raise ValueError(f"شکاف داده در MonoGrid: بلوک {i}، ردیف {j}، ستون {k}")
        return fv

    v000 = get_val(i0, j0, k0)
    v100 = get_val(i1, j0, k0)
    v010 = get_val(i0, j1, k0)
    v110 = get_val(i1, j1, k0)
    v001 = get_val(i0, j0, k1)
    v101 = get_val(i1, j0, k1)
    v011 = get_val(i0, j1, k1)
    v111 = get_val(i1, j1, k1)

    x0, x1, x = ffa_axis[i0], ffa_axis[i1], ffa
    y0, y1, y = ton_axis[j0], ton_axis[j1], ton
    z0, z1, z = hour_axis[k0], hour_axis[k1], target_hours

    v00 = lerp(x, x0, x1, v000, v100)
    v10 = lerp(x, x0, x1, v010, v110)
    v01 = lerp(x, x0, x1, v001, v101)
    v11 = lerp(x, x0, x1, v011, v111)

    v0 = lerp(y, y0, y1, v00, v10)
    v1 = lerp(y, y0, y1, v01, v11)

    return lerp(z, z0, z1, v0, v1)

def calc_lecithin(site: str, ffa: float, ton: float, hours: float, expander: Optional[str], line_mode: Optional[str]) -> float:
    grid = PRED.sheets.get("MonoGrid")
    if not grid:
        raise ValueError("شیت مرجع MonoGrid یافت نشد.")

    base24 = trilinear_interpolate(grid, ffa, ton, 24)

    if site == "Semnan":
        result24 = base24
        if ffa > 1.7:
            sem = PRED.sheets.get("semnan")
            if sem and len(sem) >= 10 and len(sem[9]) >= 2:
                try:
                    sensitivity = float(sem[9][1])  # B10
                except Exception:
                    sensitivity = float("nan")
                if not math.isnan(sensitivity):
                    ffa_axis = []
                    for v in grid[0][4:]:
                        try:
                            ffa_axis.append(float(v))
                        except Exception:
                            break
                    if ffa_axis:
                        baseline = trilinear_interpolate(grid, ffa_axis[0], ton, 24)
                        result24 = baseline + sensitivity * (base24 - baseline)
        return result24 * (hours / 24.0)

    if site == "Kermanshah":
        # latest: no-expander uses B6 = E41*0.68 => constant 0.68 modifier to base24
        if expander == "No":
            return (base24 * 0.68) * (hours / 24.0)

        # With expander:
        if line_mode == "CanolaSoya":
            sheet = PRED.sheets.get("Kermanshah with expander ")
            if not sheet:
                raise ValueError('شیت "Kermanshah with expander " یافت نشد.')
            # B8 => row 8 col B => index [7][1]
            try:
                factor = float(sheet[7][1])
            except Exception:
                raise ValueError("ضریب اصلاحی B8 معتبر نیست.")
            return (base24 * factor) * (hours / 24.0)

        return base24 * (hours / 24.0)

    raise ValueError("سایت نامعتبر است.")

# ---------------------------
# Shift performance
# ---------------------------
def moisture_comment(m: float) -> str:
    if 40 <= m <= 50:
        return "⭐ رطوبت عالی (بین 40 تا 50)"
    if 40 <= m <= 60:
        return "✅ رطوبت در رنج (بین 40 تا 60)"
    if m < 40:
        return "⚠️ رطوبت کمتر از رنج (کمتر از 40)"
    return "⚠️ رطوبت بیشتر از رنج (بیشتر از 60)"

def compute_shift_metrics(barrels: float, moisture: float, ffa: float) -> Dict[str, float]:
    if not (0 < moisture < 100):
        raise ValueError("درصد رطوبت نامعتبر است (باید بین 0 و 100 باشد).")
    lecithin_kg = barrels * 200.0
    gum_kg = lecithin_kg * 100.0 / (100.0 - moisture)
    gum_per_hour = gum_kg / 24.0
    gum_per_min = gum_kg / 1440.0
    score = gum_per_min / ffa if ffa and ffa > 0 else float("nan")
    return {
        "lecithinKg": lecithin_kg,
        "gumKg": gum_kg,
        "gumKgPerHour": gum_per_hour,
        "gumKgPerMin": gum_per_min,
        "score": score,
    }

def recompute_best_shift_for_day(data: Dict[str, Any], day: int) -> None:
    # Determine best shift by max score for that day
    day_key = str(day)
    shifts = data.get(SHIFT_KEY, {}).get(day_key, {})
    best_shift = None
    best_score = -1e18
    for sh in ["1", "2", "3"]:
        rec = shifts.get(sh)
        if not rec:
            continue
        score = rec.get("score")
        try:
            score_f = float(score)
        except Exception:
            continue
        if math.isnan(score_f):
            continue
        if score_f > best_score:
            best_score = score_f
            best_shift = sh
    # annotate all shifts
    for sh in ["1", "2", "3"]:
        rec = shifts.get(sh)
        if rec:
            rec["bestShift"] = f"شیفت {best_shift}" if best_shift == sh else "—"
    data.setdefault(SHIFT_KEY, {})[day_key] = shifts

# ---------------------------
# Telegram conversation states
# ---------------------------
(
    MAIN_MENU,
    LECITHIN_DAY, LECITHIN_SHIFT, LECITHIN_SITE, LECITHIN_EXPANDER, LECITHIN_LINE,
    LECITHIN_FFA, LECITHIN_TON, LECITHIN_HOURS, LECITHIN_SAVE_CONFIRM,

    SHIFT_DAY, SHIFT_SHIFT, SHIFT_SOURCE, SHIFT_SITE, SHIFT_EXPANDER, SHIFT_LINE,
    SHIFT_FFA, SHIFT_TON, SHIFT_HOURS, SHIFT_MOISTURE, SHIFT_BARRELS_MANUAL, SHIFT_SAVE_CONFIRM
) = range(22)

def kb(rows: List[List[Tuple[str, str]]]) -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup([
        [InlineKeyboardButton(text, callback_data=cb) for text, cb in row]
        for row in rows
    ])


async def show_main_menu(message, *, text_prefix: str = "سلام! یکی از بخش‌ها را انتخاب کنید:") -> None:
    await message.reply_text(
        text_prefix,
        reply_markup=kb([
            [("🧪 پیش‌بینی روزانه لسیتین", "menu_lecithin"), ("📊 ارزیابی عملکرد کارکنان", "menu_shift")],
            [("📤 خروجی لسیتین (Excel)", "export_lecithin"), ("📤 خروجی ارزیابی (Excel)", "export_shifts")]
        ])
    )

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    await show_main_menu(update.message)
    return MAIN_MENU

# ---------------------------
# Export handlers
# ---------------------------
def _write_csv(out_path: str, rows: list, fieldnames: list) -> None:
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    with open(out_path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        w.writeheader()
        for r in rows:
            w.writerow(r)

def _write_xlsx(out_path: str, rows: list, fieldnames: list) -> None:
    # Lightweight Excel writer using openpyxl (no pandas).
    from openpyxl import Workbook
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "data"
    ws.append(fieldnames)
    for r in rows:
        ws.append([r.get(k) for k in fieldnames])
    wb.save(out_path)


def _sort_day_shift(rows: list) -> list:
    def to_int(x):
        try:
            return int(str(x))
        except Exception:
            return 0
    return sorted(rows, key=lambda r: (to_int(r.get("Day")), to_int(r.get("Shift"))))

async def export_lecithin(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    chat_id = update.effective_chat.id
    data = load_user_data(chat_id).get(LECITHIN_KEY, {})
    rows = []
    for day, shifts in data.items():
        for sh, rec in shifts.items():
            barrels = rec.get("barrels")
            ton = rec.get("ton")
            lec_kg = (barrels * 200) if barrels is not None else None
            rows.append({
                "Day": day,
                "Shift": sh,
                "Site": rec.get("site"),
                "FFA": rec.get("ffa"),
                "OilTon": ton,
                "Hours": rec.get("hours"),
                "Expander": rec.get("expander"),
                "LineMode": rec.get("lineMode"),
                "LecithinBarrels": barrels,
                "LecithinKg": lec_kg,
                "KgPerTon": (lec_kg / ton) if (lec_kg is not None and ton) else None,
            })
    rows = _sort_day_shift(rows)
    fieldnames = ["Day","Shift","Site","FFA","OilTon","Hours","Expander","LineMode","LecithinBarrels","LecithinKg","KgPerTon"]
    out_path = os.path.join(DATA_DIR, f"lecithin_{chat_id}.xlsx")
    _write_xlsx(out_path, rows, fieldnames)

    # Send as a Telegram document so it can be opened on phone
    try:
        if update.callback_query:
            await update.callback_query.answer()
        chat_id2 = update.effective_chat.id
        with open(out_path, "rb") as f:
            await context.bot.send_document(chat_id=chat_id2, document=f, filename="lecithin_export.xlsx")
    except Exception as e:
        msg = f"خطا در ارسال فایل اکسل: {e}"
        if update.callback_query and update.callback_query.message:
            await update.callback_query.message.reply_text(msg)
        elif update.message:
            await update.message.reply_text(msg)

async def export_shifts(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    chat_id = update.effective_chat.id
    data = load_user_data(chat_id).get(SHIFT_KEY, {})
    rows = []
    for day, shifts in data.items():
        for sh, rec in shifts.items():
            rows.append({
                "Day": day,
                "Shift": sh,
                "FFA": rec.get("ffa"),
                "OilTon": rec.get("ton"),
                "Hours": rec.get("hours"),
                "Moisture": rec.get("moisture"),
                "LecithinBarrels": rec.get("barrels"),
                "LecithinKg": rec.get("lecithinKg"),
                "GumKgDaily": rec.get("gumKgDaily"),
                "GumKgPerHour": rec.get("gumKgPerHour"),
                "GumKgPerMin": rec.get("gumKgPerMin"),
                "MoistureStatus": rec.get("moistureStatus"),
                "Score(gum_per_min/FFA)": rec.get("score"),
                "BestShift": rec.get("bestShift"),
            })
    rows = _sort_day_shift(rows)
    fieldnames = ["Day","Shift","FFA","OilTon","Hours","Moisture","LecithinBarrels","LecithinKg","GumKgDaily","GumKgPerHour","GumKgPerMin","MoistureStatus","Score(gum_per_min/FFA)","BestShift"]
    out_path = os.path.join(DATA_DIR, f"shifts_{chat_id}.xlsx")
    _write_xlsx(out_path, rows, fieldnames)

    # Send as a Telegram document so it can be opened on phone
    try:
        if update.callback_query:
            await update.callback_query.answer()
        chat_id2 = update.effective_chat.id
        with open(out_path, "rb") as f:
            await context.bot.send_document(chat_id=chat_id2, document=f, filename="shift_export.xlsx")
    except Exception as e:
        msg = f"خطا در ارسال فایل اکسل: {e}"
        if update.callback_query and update.callback_query.message:
            await update.callback_query.message.reply_text(msg)
        elif update.message:
            await update.message.reply_text(msg)

async def menu_router(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    q = update.callback_query
    await q.answer()
    if q.data == "restart":
        context.user_data.clear()
        await show_main_menu(q.message, text_prefix="🔄 از نو شروع شد. یکی از بخش‌ها را انتخاب کنید:")
        return MAIN_MENU
    if q.data == "back_main":
        await show_main_menu(q.message)
        return MAIN_MENU
    if q.data == "menu_lecithin":
        # Start by collecting inputs first, then ask which day/shift to register.
        context.user_data.clear()
        await q.message.reply_text("سایت را انتخاب کنید:", reply_markup=kb([[("سمنان", "lec_site_Semnan"), ("کرمانشاه", "lec_site_Kermanshah")]]))
        return LECITHIN_SITE
    if q.data == "menu_shift":
        context.user_data.clear()
        await q.message.reply_text("سایت را انتخاب کنید:", reply_markup=kb([
            [("سمنان", "sh_site_Semnan"), ("کرمانشاه", "sh_site_Kermanshah")],
            [("⬅️ منوی اصلی", "back_main"), ("🔄 شروع مجدد", "restart")]
        ]))
        return SHIFT_SITE
    if q.data == "export_lecithin":
        await export_lecithin(update, context)
        return MAIN_MENU
    if q.data == "export_shifts":
        await export_shifts(update, context)
        return MAIN_MENU
    return MAIN_MENU


async def nav_router(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    q = update.callback_query
    await q.answer()
    data = q.data

    # Global
    if data in ("back_main", "restart"):
        return await menu_router(update, context)

    # Lecithin back steps
    if data == "lec_back_site":
        await q.message.reply_text("سایت را انتخاب کنید:", reply_markup=kb([[('سمنان', 'lec_site_Semnan'), ('کرمانشاه', 'lec_site_Kermanshah')],[('⬅️ منوی اصلی','back_main'),('🔄 شروع مجدد','restart')]]))
        return LECITHIN_SITE

    if data == "lec_back_expander":
        await q.message.reply_text("اکسپندر در مدار هست؟", reply_markup=kb([
            [("بله", "lec_exp_Yes"), ("خیر", "lec_exp_No")],
            [("⬅️ مرحله قبل", "lec_back_site"), ("🔄 شروع مجدد", "restart")],
            [("⬅️ منوی اصلی", "back_main")]
        ]))
        return LECITHIN_EXPANDER

    if data == "lec_back_line":
        await q.message.reply_text("خط را انتخاب کنید:", reply_markup=kb([
            [("نرمال", "lec_line_Normal"), ("کلزا/سویا", "lec_line_CanolaSoya")],
            [("⬅️ مرحله قبل", "lec_back_expander"), ("🔄 شروع مجدد", "restart")],
            [("⬅️ منوی اصلی", "back_main")]
        ]))
        return LECITHIN_LINE

    if data == "lec_back_ffa":
        await q.message.reply_text("FFA را وارد کنید (مثلاً 1.8):", reply_markup=kb([[('⬅️ مرحله قبل','lec_back_line'),('🔄 شروع مجدد','restart')],[('⬅️ منوی اصلی','back_main')]]))
        return LECITHIN_FFA

    if data == "lec_back_ton":
        await q.message.reply_text("🛢 تناژ روغن را وارد کنید (مثلاً 120):", reply_markup=kb([[('⬅️ مرحله قبل','lec_back_ffa'),('🔄 شروع مجدد','restart')],[('⬅️ منوی اصلی','back_main')]]))
        return LECITHIN_TON

    if data == "lec_back_hours":
        await q.message.reply_text("ساعات تولید را وارد کنید (مثلاً 8):",
                                   reply_markup=kb([[("🔄 شروع مجدد", "restart"), ("⬅️ منوی اصلی", "back_main")]]))
        return LECITHIN_HOURS

    # Shift back steps
    if data == "sh_back_site":
        await q.message.reply_text("سایت را انتخاب کنید:", reply_markup=kb([
            [("سمنان", "sh_site_Semnan"), ("کرمانشاه", "sh_site_Kermanshah")],
            [("⬅️ منوی اصلی", "back_main"), ("🔄 شروع مجدد", "restart")]
        ]))
        return SHIFT_SITE
    if data == "sh_back_day":
        # show day picker again
        await q.message.reply_text("روز را انتخاب کنید:", reply_markup=kb([
            [(f"روز {i}", f"sh_day_{i}") for i in range(1,6)],
            [(f"روز {i}", f"sh_day_{i}") for i in range(6,11)],
            [(f"روز {i}", f"sh_day_{i}") for i in range(11,16)],
            [(f"روز {i}", f"sh_day_{i}") for i in range(16,21)],
            [(f"روز {i}", f"sh_day_{i}") for i in range(21,26)],
            [(f"روز {i}", f"sh_day_{i}") for i in range(26,32)],
            [("⬅️ مرحله قبل", "sh_back_site"), ("🔄 شروع مجدد", "restart")],
            [("⬅️ منوی اصلی", "back_main")]
        ]))
        return SHIFT_DAY
    if data == "sh_back_shift":
        # show shift picker again
        await q.message.reply_text("شیفت را انتخاب کنید:", reply_markup=kb([
            [("شیفت 1", "sh_shift_1"), ("شیفت 2", "sh_shift_2"), ("شیفت 3", "sh_shift_3")],
            [("⬅️ مرحله قبل", "sh_back_day"), ("🔄 شروع مجدد", "restart")],
            [("⬅️ منوی اصلی", "back_main")]
        ]))
        return SHIFT_SHIFT

    return MAIN_MENU

# ---------------------------
# Lecithin flow
# ---------------------------
async def lecithin_day(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    q = update.callback_query
    await q.answer()
    day = int(q.data.split("_")[-1])
    context.user_data["lec_day"] = day

    pending = context.user_data.get("pending_lecithin")
    if not pending:
        await q.message.reply_text("اطلاعات محاسبه پیدا نشد. لطفاً دوباره از /start شروع کنید.", reply_markup=kb([[("⬅️ منوی اصلی", "back_main")]]))
        return MAIN_MENU

    # Save immediately (بدون شیفت)
    chat_id = update.effective_chat.id
    user_data = load_user_data(chat_id)
    lec = user_data.get(LECITHIN_KEY, {})
    day_key = str(day)
    sh_key = ""  # بدون شیفت
    lec.setdefault(day_key, {})
    lec[day_key][sh_key] = {
        "site": pending.get("site"),
        "expander": pending.get("expander"),
        "lineMode": pending.get("lineMode"),
        "ffa": pending.get("ffa"),
        "ton": pending.get("ton"),
        "hours": pending.get("hours"),
        "barrels": pending.get("barrels"),
    }
    user_data[LECITHIN_KEY] = lec
    save_user_data(chat_id, user_data)

    barrels = float(pending.get("barrels") or 0.0)
    ton = float(pending.get("ton") or 0.0)
    kg = barrels * 200.0
    kg_per_ton = (kg / ton) if ton else 0.0

    await q.message.reply_text(
        f"✅ ثبت شد (روز {day})\n\n"
        f"لسیتین: {barrels:.3f} بشکه | {kg:.1f} کیلوگرم | {kg_per_ton:.2f} کیلوگرم/تن\n\n"
        f"برای شروع مجدد از همین بخش می‌توانید /start را بزنید.",
        reply_markup=kb([[("⬅️ منوی اصلی", "back_main"), ("🔄 شروع مجدد", "restart")]])
    )
    context.user_data.pop("pending_lecithin", None)
    return MAIN_MENU



async def lecithin_shift(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    q = update.callback_query
    await q.answer()
    sh = q.data.split("_")[-1]
    day = context.user_data.get("lec_day")
    pending = context.user_data.get("pending_lecithin")

    if not pending or day is None:
        await q.message.reply_text("اطلاعات محاسبه پیدا نشد. لطفاً دوباره از /start شروع کنید.")
        return MAIN_MENU

    # Save immediately
    chat_id = update.effective_chat.id
    user_data = load_user_data(chat_id)
    lec = user_data.get(LECITHIN_KEY, {})
    day_key = str(day)
    sh_key = str(sh)
    lec.setdefault(day_key, {})
    lec[day_key][sh_key] = {
        "site": pending.get("site"),
        "expander": pending.get("expander"),
        "lineMode": pending.get("lineMode"),
        "ffa": pending.get("ffa"),
        "ton": pending.get("ton"),
        "hours": pending.get("hours"),
        "barrels": pending.get("barrels"),
    }
    user_data[LECITHIN_KEY] = lec
    save_user_data(chat_id, user_data)

    barrels = float(pending.get("barrels") or 0.0)
    ton = float(pending.get("ton") or 0.0)
    kg = barrels * 200.0
    kg_per_ton = (kg / ton) if ton else 0.0

    await q.message.reply_text(
        f"✅ ثبت شد (روز {day} - شیفت {sh})\n\n"
        f"لسیتین: {barrels:.3f} بشکه | {kg:.1f} کیلوگرم | {kg_per_ton:.2f} کیلوگرم/تن\n\n"
        f"اگر خروجی اکسل می‌خواهید، از منوی اصلی «📤 خروجی لسیتین (Excel)» را بزنید.",
        reply_markup=kb([[("⬅️ منوی اصلی", "back_main")]])
    )
    # clean pending
    context.user_data.pop("pending_lecithin", None)
    return MAIN_MENU


async def lecithin_site(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    q = update.callback_query
    await q.answer()
    site = q.data.split("_")[-1]
    context.user_data["site"] = site

    # Kermanshah has extra options (expander + line mode)
    if site == "Kermanshah":
        await q.message.reply_text(
            "🔧 اکسپندر در مدار هست؟\n\nلطفاً وضعیت اکسپندر را مشخص کنید:",
            reply_markup=kb([[("✅ بله", "lec_exp_Yes"), ("❌ خیر", "lec_exp_No")]]),
        )
        return LECITHIN_EXPANDER

    # Semnan: no expander / line mode step
    context.user_data["expander"] = None
    context.user_data["lineMode"] = None
    await q.message.reply_text("FFA را وارد کنید (مثلاً 1.8):", reply_markup=kb([[("⬅️ مرحله قبل", "lec_back_site"), ("🔄 شروع مجدد", "restart")],[("⬅️ منوی اصلی", "back_main")]]))
    return LECITHIN_FFA

async def lecithin_expander(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    q = update.callback_query
    await q.answer()
    exp = q.data.split("_")[-1]
    context.user_data["expander"] = exp
    await q.message.reply_text("🕹حالت خط را انتخاب کنید:", reply_markup=kb([[("نرمال", "lec_line_Normal"), ("کلزا-سویا", "lec_line_CanolaSoya")]]))
    return LECITHIN_LINE

async def lecithin_line(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    q = update.callback_query
    await q.answer()
    line = q.data.split("_")[-1]
    context.user_data["lineMode"] = line
    await q.message.reply_text("FFA را وارد کنید (مثلاً 1.8):", reply_markup=kb([[("⬅️ مرحله قبل", "lec_back_site"), ("🔄 شروع مجدد", "restart")],[("⬅️ منوی اصلی", "back_main")]]))
    return LECITHIN_FFA

async def lecithin_ffa(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    try:
        ffa = float(update.message.text.strip())
    except Exception:
        await update.message.reply_text("FFA نامعتبر است. دوباره وارد کنید:")
        return LECITHIN_FFA
    context.user_data["ffa"] = ffa
    await update.message.reply_text("🛢 تناژ روغن را وارد کنید (مثلاً 120):", reply_markup=kb([[("⬅️ مرحله قبل", "lec_back_ffa"), ("🔄 شروع مجدد", "restart")],[("⬅️ منوی اصلی", "back_main")]]))
    return LECITHIN_TON

async def lecithin_ton(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    try:
        ton = float(update.message.text.strip())
    except Exception:
        await update.message.reply_text("تناژ نامعتبر است. دوباره وارد کنید:")
        return LECITHIN_TON
    context.user_data["ton"] = ton
    await update.message.reply_text("ساعات تولید را وارد کنید (مثلاً 20):")
    return LECITHIN_HOURS

async def lecithin_hours(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    try:
        hours = float(update.message.text.strip())

        # اصلاح ساعات برای سایت سمنان
        site = context.user_data.get("site")
        if site == "Semnan" and hours < 24:
            difference = 24 - hours
            bonus = difference / 2
            hours = hours + bonus
    except Exception:
        await update.message.reply_text("ساعت نامعتبر است. دوباره وارد کنید:")
        return LECITHIN_HOURS
    context.user_data["hours"] = hours

    site = context.user_data["site"]
    exp = context.user_data.get("expander")
    line = context.user_data.get("lineMode")
    ffa = context.user_data["ffa"]
    ton = context.user_data["ton"]

    try:
        barrels = calc_lecithin(site, ffa, ton, hours, exp, line)
    except Exception as e:
        await update.message.reply_text(f"خطا در محاسبه: {e}")
        return ConversationHandler.END

    kg = barrels * 200.0
    kg_per_ton = (kg / ton) if ton else float("nan")

        # store pending result, then ask user which day/shift to register
    context.user_data["pending_lecithin"] = {
        "site": site,
        "expander": exp,
        "lineMode": line,
        "ffa": ffa,
        "ton": ton,
        "hours": hours,
        "barrels": barrels,
    }

    msg = (
        f"🧾 🧾 نتیجه لسیتین روزانه پیش‌بینی شده پیش‌بینی شده\n"
        f"- لسیتین: <b>{barrels:.3f}</b> بشکه\n"
        f"- لسیتین: <b>{kg:.1f}</b> کیلوگرم\n"
        f"- نسبت به تناژ روغن: <b>{kg_per_ton:.2f}</b> کیلوگرم/تن\n\n"
        f"برای ثبت، روز را انتخاب کنید:"
    )
    await update.message.reply_text(
        msg,
        parse_mode=ParseMode.HTML,
        reply_markup=kb([
            [(f"روز {i}", f"lec_day_{i}") for i in range(1,6)],
            [(f"روز {i}", f"lec_day_{i}") for i in range(6,11)],
            [(f"روز {i}", f"lec_day_{i}") for i in range(11,16)],
            [(f"روز {i}", f"lec_day_{i}") for i in range(16,21)],
            [(f"روز {i}", f"lec_day_{i}") for i in range(21,26)],
            [(f"روز {i}", f"lec_day_{i}") for i in range(26,32)],
        ])
    )
    return LECITHIN_DAY


async def lecithin_save_confirm(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    q = update.callback_query
    await q.answer()
    if q.data == "lec_save_no":
        await q.message.reply_text("اوکی. /start")
        return ConversationHandler.END

    chat_id = update.effective_chat.id
    data = load_user_data(chat_id)

    day = str(context.user_data["lec_day"])
    sh = str(context.user_data["lec_shift"])

    rec = {
        "site": context.user_data["site"],
        "expander": context.user_data.get("expander"),
        "lineMode": context.user_data.get("lineMode"),
        "ffa": context.user_data["ffa"],
        "ton": context.user_data["ton"],
        "hours": context.user_data["hours"],
        "barrels": context.user_data["barrels"],
    }

    data.setdefault(LECITHIN_KEY, {}).setdefault(day, {})[sh] = rec
    save_user_data(chat_id, data)
    await q.message.reply_text("✅ ذخیره شد. /start")
    return ConversationHandler.END

# ---------------------------
# Shift flow
# ---------------------------

async def shift_site(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    q = update.callback_query
    await q.answer()
    site = q.data.split("_")[-1]
    context.user_data["site"] = site
    await q.message.reply_text("روز را انتخاب کنید:", reply_markup=kb([
        [(f"روز {i}", f"sh_day_{i}") for i in range(1,6)],
        [(f"روز {i}", f"sh_day_{i}") for i in range(6,11)],
        [(f"روز {i}", f"sh_day_{i}") for i in range(11,16)],
        [(f"روز {i}", f"sh_day_{i}") for i in range(16,21)],
        [(f"روز {i}", f"sh_day_{i}") for i in range(21,26)],
        [(f"روز {i}", f"sh_day_{i}") for i in range(26,32)],
        [("⬅️ مرحله قبل", "sh_back_site"), ("🔄 شروع مجدد", "restart")],
        [("⬅️ منوی اصلی", "back_main")]
    ]))
    return SHIFT_DAY
async def shift_day(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    q = update.callback_query
    await q.answer()
    day = int(q.data.split("_")[-1])
    context.user_data["sh_day"] = day
    await q.message.reply_text("شیفت را انتخاب کنید:", reply_markup=kb([
        [("شیفت 1", "sh_shift_1"), ("شیفت 2", "sh_shift_2"), ("شیفت 3", "sh_shift_3")],
        [("⬅️ مرحله قبل", "sh_back_day"), ("🔄 شروع مجدد", "restart")],
        [("⬅️ منوی اصلی", "back_main")]
    ]))
    return SHIFT_SHIFT

async def shift_shift(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    q = update.callback_query
    await q.answer()
    sh = q.data.split("_")[-1]
    context.user_data["sh_shift"] = sh

    await q.message.reply_text("لسیتین (بشکه) از کجا بیاد؟", reply_markup=kb([
        [("از «پیش‌بینی روزانه لسیتین»", "sh_src_from_lec"), ("ورودی دسته بشکه (فقط تعداد بشکه)", "sh_src_manual_only")],
        [("⬅️ مرحله قبل", "sh_back_shift"), ("🔄 شروع مجدد", "restart")],
        [("⬅️ منوی اصلی", "back_main")]
    ]))
    return SHIFT_SOURCE

async def shift_source(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    q = update.callback_query
    await q.answer()
    src = q.data.replace("sh_src_", "")  # from_lec | manual_only
    context.user_data["sh_src"] = src

    if src == "from_lec":
        chat_id = update.effective_chat.id
        data = load_user_data(chat_id).get(LECITHIN_KEY, {})
        day_key = str(context.user_data.get("sh_day"))

        day_rec = data.get(day_key, {})
        rec = None
        # Prefer daily record (بدون شیفت)
        if "" in day_rec:
            rec = day_rec.get("")
        elif day_rec:
            # take first available
            rec = next(iter(day_rec.values()))
        if not rec:
            await q.message.reply_text(
                "برای این روز در بخش «پیش‌بینی روزانه لسیتین» داده‌ای ثبت نشده. گزینه «ورودی دسته بشکه» را انتخاب کنید.",
                reply_markup=kb([[("ورودی دسته بشکه (فقط تعداد بشکه)", "sh_src_manual_only")],
                                 [("⬅️ مرحله قبل", "sh_back_shift"), ("🔄 شروع مجدد", "restart")],
                                 [("⬅️ منوی اصلی", "back_main")]])
            )
            return SHIFT_SOURCE

        context.user_data["barrels"] = float(rec.get("barrels") or 0.0)
        # اگر در ثبت «پیش‌بینی روزانه لسیتین» موجود باشد، برای گزارش ذخیره می‌کنیم
        try:
            context.user_data["ton"] = float(rec.get("ton")) if rec.get("ton") not in (None, "") else None
        except Exception:
            context.user_data["ton"] = None
        try:
            context.user_data["hours"] = float(rec.get("hours")) if rec.get("hours") not in (None, "") else None
        except Exception:
            context.user_data["hours"] = None
        # اگر ffa موجود بود برای امتیاز استفاده می‌کنیم
        ffa_val = rec.get("ffa")
        try:
            context.user_data["ffa"] = float(ffa_val) if ffa_val is not None else None
        except Exception:
            context.user_data["ffa"] = None

        await q.message.reply_text(
            "درصد رطوبت گام را وارد کنید (مثلاً 45):",
            reply_markup=kb([[("⬅️ مرحله قبل", "sh_back_shift"), ("🔄 شروع مجدد", "restart")],
                             [("⬅️ منوی اصلی", "back_main")]])
        )
        return SHIFT_MOISTURE

    # manual_only: فقط تعداد بشکه
    await q.message.reply_text(
        "تعداد بشکه لسیتین را وارد کنید (مثلاً 44.93):",
        reply_markup=kb([[("⬅️ مرحله قبل", "sh_back_shift"), ("🔄 شروع مجدد", "restart")],
                         [("⬅️ منوی اصلی", "back_main")]])
    )
    # ffa را در این حالت نداریم
    context.user_data["ffa"] = None
    context.user_data["ton"] = None
    context.user_data["hours"] = None
    return SHIFT_BARRELS_MANUAL



async def shift_ffa(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    try:
        ffa = float(update.message.text.strip())
    except Exception:
        await update.message.reply_text("FFA نامعتبر است. دوباره وارد کنید:")
        return SHIFT_FFA
    context.user_data["ffa"] = ffa
    await update.message.reply_text("🛢 تناژ روغن را وارد کنید (مثلاً 120):", reply_markup=kb([[("⬅️ مرحله قبل", "lec_back_ffa"), ("🔄 شروع مجدد", "restart")],[("⬅️ منوی اصلی", "back_main")]]))
    return SHIFT_TON

async def shift_ton(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    try:
        ton = float(update.message.text.strip())
    except Exception:
        await update.message.reply_text("تناژ نامعتبر است. دوباره وارد کنید:")
        return SHIFT_TON
    context.user_data["ton"] = ton
    await update.message.reply_text("ساعات تولید را وارد کنید (مثلاً 8):", reply_markup=kb([[("⬅️ مرحله قبل", "lec_back_ton"), ("🔄 شروع مجدد", "restart")],[("⬅️ منوی اصلی", "back_main")]]))
    return SHIFT_HOURS

async def shift_hours(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    try:
        hours = float(update.message.text.strip())
    except Exception:
        await update.message.reply_text("ساعت نامعتبر است. دوباره وارد کنید:")
        return SHIFT_HOURS
    context.user_data["hours"] = hours
    await update.message.reply_text("لسیتین تولیدی (بشکه) را وارد کنید (مثلاً 44.93):")
    return SHIFT_BARRELS_MANUAL

async def shift_barrels_manual(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    try:
        barrels = float(update.message.text.strip())
    except Exception:
        await update.message.reply_text("عدد بشکه نامعتبر است. دوباره وارد کنید:")
        return SHIFT_BARRELS_MANUAL
    context.user_data["barrels"] = barrels
    await update.message.reply_text("درصد رطوبت گام را وارد کنید (مثلاً 45):")
    return SHIFT_MOISTURE

async def shift_moisture(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    try:
        moisture = float(update.message.text.strip())
    except Exception:
        await update.message.reply_text("درصد رطوبت نامعتبر است. دوباره وارد کنید:")
        return SHIFT_MOISTURE

    context.user_data["moisture"] = moisture
    ffa_raw = context.user_data.get("ffa")
    ffa = float(ffa_raw) if ffa_raw not in (None, "") else 0.0
    barrels = float(context.user_data["barrels"])

    try:
        metrics = compute_shift_metrics(barrels, moisture, ffa)
    except Exception as e:
        await update.message.reply_text(f"خطا: {e}\nدوباره درصد رطوبت را وارد کنید:")
        return SHIFT_MOISTURE

    status = moisture_comment(moisture)

    score_line = (
        f"- امتیاز (گام/دقیقه ÷ FFA): <b>{metrics['score']:.4f}</b>" if not math.isnan(metrics['score'])
        else "- امتیاز: <b>-</b>"
    )

    day = context.user_data["sh_day"]
    sh = context.user_data["sh_shift"]

    msg = (
        f"👷 عملکرد شیفت (روز {day} - شیفت {sh})\n"
        f"- لسیتین: <b>{barrels:.3f}</b> بشکه\n"
        f"- لسیتین: <b>{metrics['lecithinKg']:.1f}</b> کیلوگرم\n"
        f"- وزن گام: <b>{metrics['gumKg']:.1f}</b> کیلوگرم\n"
        f"- گام/ساعت: <b>{metrics['gumKgPerHour']:.2f}</b> kg/h\n"
        f"- گام/دقیقه: <b>{metrics['gumKgPerMin']:.3f}</b> kg/min\n"
        f"- وضعیت رطوبت: {status}\n"
        f"{score_line}\n\n"
        f"ذخیره شود؟"
    )
    context.user_data["metrics"] = metrics = metrics
    context.user_data["moistureStatus"] = status

    await update.message.reply_text(msg, parse_mode=ParseMode.HTML,
                                   reply_markup=kb([[("💾 ذخیره", "sh_save_yes"), ("❌ نه", "sh_save_no")],
                                                 [("🔄 شروع مجدد", "restart"), ("⬅️ منوی اصلی", "back_main")]]))
    return SHIFT_SAVE_CONFIRM

async def shift_save_confirm(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    q = update.callback_query
    await q.answer()
    if q.data == "sh_save_no":
        await q.message.reply_text("اوکی. /start")
        return ConversationHandler.END

    chat_id = update.effective_chat.id
    all_data = load_user_data(chat_id)

    day = str(context.user_data["sh_day"])
    sh = str(context.user_data["sh_shift"])

    metrics = context.user_data["metrics"]

    rec = {
        "site": context.user_data.get("site"),
        "source": context.user_data.get("sh_src"),
        "ffa": context.user_data.get("ffa"),
        "ton": context.user_data.get("ton"),
        "hours": context.user_data.get("hours"),
        "barrels": context.user_data.get("barrels"),
        "moisture": context.user_data.get("moisture"),
        "moistureStatus": context.user_data.get("moistureStatus"),
        "lecithinKg": metrics.get("lecithinKg"),
        "gumKgDaily": metrics.get("gumKg"),
        "gumKgPerHour": metrics.get("gumKgPerHour"),
        "gumKgPerMin": metrics.get("gumKgPerMin"),
        "score": metrics.get("score"),
        "bestShift": "—",
    }

    all_data.setdefault(SHIFT_KEY, {}).setdefault(day, {})[sh] = rec
    recompute_best_shift_for_day(all_data, int(day))
    save_user_data(chat_id, all_data)

    # Inform best shift for day if available
    best = None
    shifts = all_data.get(SHIFT_KEY, {}).get(day, {})
    for s in ["1","2","3"]:
        r = shifts.get(s)
        if r and r.get("bestShift","—") != "—":
            best = r["bestShift"]
            break

    extra = f"\n🏆 بهترین شیفت این روز: {best}" if best else ""
    await q.message.reply_text(f"✅ ذخیره شد.{extra}\n/start")
    return ConversationHandler.END

# ---------------------------
# Fallback / cancel
# ---------------------------
async def cancel(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    await update.message.reply_text("لغو شد. /start")
    return ConversationHandler.END

# ---------------------------
# App bootstrap
# ---------------------------
def main() -> None:
    load_dotenv(os.path.join(BASE_DIR, ".env"))
    token = os.getenv("BOT_TOKEN", "").strip()
    if not token:
        raise RuntimeError("BOT_TOKEN در فایل .env تنظیم نشده است.")

    app = Application.builder().token(token).build()

    # Exports (commands + callbacks)
    app.add_handler(CommandHandler("export_lecithin", export_lecithin))
    app.add_handler(CommandHandler("export_shifts", export_shifts))

    # Admin utilities
    app.add_handler(CommandHandler("myid", myid))
    app.add_handler(CommandHandler("admin_stats", admin_stats))

    # Conversation handler
    conv = ConversationHandler(
        entry_points=[CommandHandler("start", start)],
        states={
            MAIN_MENU: [CallbackQueryHandler(menu_router)],

            LECITHIN_DAY: [CallbackQueryHandler(nav_router, pattern=r'^(back_main|restart|lec_back_site|lec_back_expander|lec_back_line|lec_back_ffa|lec_back_ton|lec_back_hours|sh_back_site|sh_back_day|sh_back_shift)$'), CallbackQueryHandler(lecithin_day, pattern=r"^lec_day_\d+$")],
            LECITHIN_SITE: [CallbackQueryHandler(nav_router, pattern=r'^(back_main|restart|lec_back_site|lec_back_expander|lec_back_line|lec_back_ffa|lec_back_ton|lec_back_hours|sh_back_site|sh_back_day|sh_back_shift)$'), CallbackQueryHandler(lecithin_site, pattern=r"^lec_site_(Semnan|Kermanshah)$")],
            LECITHIN_EXPANDER: [CallbackQueryHandler(nav_router, pattern=r'^(back_main|restart|lec_back_site|lec_back_expander|lec_back_line|lec_back_ffa|lec_back_ton|lec_back_hours|sh_back_site|sh_back_day|sh_back_shift)$'), CallbackQueryHandler(lecithin_expander, pattern=r"^lec_exp_(Yes|No)$")],
            LECITHIN_LINE: [CallbackQueryHandler(nav_router, pattern=r'^(back_main|restart|lec_back_site|lec_back_expander|lec_back_line|lec_back_ffa|lec_back_ton|lec_back_hours|sh_back_site|sh_back_day|sh_back_shift)$'), CallbackQueryHandler(lecithin_line, pattern=r"^lec_line_(Normal|CanolaSoya)$")],
            LECITHIN_FFA: [MessageHandler(filters.TEXT & ~filters.COMMAND, lecithin_ffa)],
            LECITHIN_TON: [MessageHandler(filters.TEXT & ~filters.COMMAND, lecithin_ton)],
            LECITHIN_HOURS: [CallbackQueryHandler(nav_router, pattern=r'^(back_main|restart|lec_back_site|lec_back_expander|lec_back_line|lec_back_ffa|lec_back_ton|lec_back_hours)$'), MessageHandler(filters.TEXT & ~filters.COMMAND, lecithin_hours)],
            LECITHIN_SAVE_CONFIRM: [CallbackQueryHandler(lecithin_save_confirm, pattern=r"^lec_save_(yes|no)$")],

            SHIFT_SITE: [CallbackQueryHandler(nav_router, pattern=r'^(back_main|restart|lec_back_site|lec_back_expander|lec_back_line|lec_back_ffa|lec_back_ton|lec_back_hours|sh_back_site|sh_back_day|sh_back_shift)$'), CallbackQueryHandler(shift_site, pattern=r"^sh_site_(Semnan|Kermanshah)$")],

            SHIFT_DAY: [CallbackQueryHandler(nav_router, pattern=r'^(back_main|restart|lec_back_site|lec_back_expander|lec_back_line|lec_back_ffa|lec_back_ton|lec_back_hours|sh_back_site|sh_back_day|sh_back_shift)$'), CallbackQueryHandler(shift_day, pattern=r"^sh_day_\d+$")],
            SHIFT_SHIFT: [CallbackQueryHandler(nav_router, pattern=r'^(back_main|restart|lec_back_site|lec_back_expander|lec_back_line|lec_back_ffa|lec_back_ton|lec_back_hours|sh_back_site|sh_back_day|sh_back_shift)$'), CallbackQueryHandler(shift_shift, pattern=r"^sh_shift_[123]$")],
            SHIFT_SOURCE: [CallbackQueryHandler(nav_router, pattern=r'^(back_main|restart|lec_back_site|lec_back_expander|lec_back_line|lec_back_ffa|lec_back_ton|lec_back_hours|sh_back_site|sh_back_day|sh_back_shift)$'), CallbackQueryHandler(shift_source, pattern=r"^sh_src_(from_lec|manual_only)$")],
            SHIFT_FFA: [MessageHandler(filters.TEXT & ~filters.COMMAND, shift_ffa)],
            SHIFT_TON: [MessageHandler(filters.TEXT & ~filters.COMMAND, shift_ton)],
            SHIFT_HOURS: [MessageHandler(filters.TEXT & ~filters.COMMAND, shift_hours)],
            SHIFT_BARRELS_MANUAL: [CallbackQueryHandler(nav_router, pattern=r'^(back_main|restart|sh_back_shift)$'), MessageHandler(filters.TEXT & ~filters.COMMAND, shift_barrels_manual)],
            SHIFT_MOISTURE: [CallbackQueryHandler(nav_router, pattern=r'^(back_main|restart|sh_back_shift|sh_back_day|sh_back_site)$'), MessageHandler(filters.TEXT & ~filters.COMMAND, shift_moisture)],
            SHIFT_SAVE_CONFIRM: [CallbackQueryHandler(nav_router, pattern=r'^(back_main|restart|lec_back_site|lec_back_expander|lec_back_line|lec_back_ffa|lec_back_ton|lec_back_hours|sh_back_site|sh_back_day|sh_back_shift)$'), CallbackQueryHandler(shift_save_confirm, pattern=r"^sh_save_(yes|no)$")],
        },
        fallbacks=[CommandHandler("cancel", cancel)],
        allow_reentry=True,
    )
    app.add_handler(conv)

    # If user presses export in main menu callbacks
    app.add_handler(CallbackQueryHandler(export_lecithin, pattern=r"^export_lecithin$"))
    app.add_handler(CallbackQueryHandler(export_shifts, pattern=r"^export_shifts$"))
    # --- Windows/Python 3.12+ event loop fix (Python 3.14 raises if no loop set) ---
    import asyncio
    import sys as _sys
    if _sys.platform.startswith('win'):
        try:
            asyncio.set_event_loop_policy(asyncio.WindowsSelectorEventLoopPolicy())
        except Exception:
            pass
    try:
        asyncio.get_event_loop()
    except RuntimeError:
        asyncio.set_event_loop(asyncio.new_event_loop())

    app.run_polling(close_loop=False)

if __name__ == "__main__":
    main()